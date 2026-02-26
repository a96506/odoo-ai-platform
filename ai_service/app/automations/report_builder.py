"""
Natural Language Report Builder — AI-powered report generation from Odoo data.

Users type natural language queries like "Sales by product category for Q4 2025",
AI parses them into structured Odoo queries, fetches data, and formats results.
Supports Excel/PDF export and scheduled reports.
"""

import io
import os
import tempfile
from datetime import datetime, date, timedelta
from typing import Any

import structlog

from app.automations.base import BaseAutomation
from app.models.schemas import AutomationResult

logger = structlog.get_logger()

REPORT_DIR = os.path.join(tempfile.gettempdir(), "odoo_reports")

REPORT_PARSE_PROMPT = """You are an Odoo ERP report query parser. Given a natural language
report request, extract the structured query parameters needed to fetch data from Odoo.

Available Odoo models and their key fields:
- sale.order: name, partner_id, amount_total, state, date_order, user_id
- sale.order.line: product_id, product_uom_qty, price_unit, price_subtotal, order_id
- account.move: name, partner_id, amount_total, state, move_type, invoice_date, payment_state
- account.move.line: product_id, quantity, price_unit, price_subtotal, move_id, account_id
- crm.lead: name, partner_id, expected_revenue, probability, stage_id, user_id, date_deadline
- product.template: name, list_price, categ_id, type, sale_ok
- product.product: name, qty_available, virtual_available, categ_id
- res.partner: name, email, phone, customer_rank, supplier_rank, country_id
- stock.picking: name, partner_id, state, scheduled_date, picking_type_id
- purchase.order: name, partner_id, amount_total, state, date_order
- hr.expense: name, employee_id, total_amount, state, date

Date context: Today is {today}. Interpret relative dates accordingly.
- "Q1 2025" = 2025-01-01 to 2025-03-31
- "Q4 2025" = 2025-10-01 to 2025-12-31
- "last quarter" = previous calendar quarter
- "this month" = current calendar month
- "YTD" = January 1 of current year to today
- "last year" = previous calendar year"""

REPORT_PARSE_TOOLS = [
    {
        "name": "odoo_report_query",
        "description": "Structured Odoo query parameters extracted from natural language",
        "input_schema": {
            "type": "object",
            "properties": {
                "model": {
                    "type": "string",
                    "description": "Primary Odoo model to query, e.g. 'sale.order'",
                },
                "fields": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Fields to retrieve",
                },
                "domain": {
                    "type": "array",
                    "description": "Odoo domain filter as list of tuples, e.g. [['state','=','sale']]",
                },
                "group_by": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Fields to group by",
                },
                "order_by": {
                    "type": "string",
                    "description": "Sort order, e.g. 'amount_total desc'",
                },
                "limit": {
                    "type": "integer",
                    "description": "Max records to return",
                },
                "title": {
                    "type": "string",
                    "description": "Human-readable report title",
                },
                "comparison": {
                    "type": "object",
                    "description": "If comparison requested, contains period info",
                    "properties": {
                        "enabled": {"type": "boolean"},
                        "compare_domain": {
                            "type": "array",
                            "description": "Domain for comparison period",
                        },
                        "compare_label": {"type": "string"},
                    },
                },
            },
            "required": ["model", "fields", "domain", "title"],
        },
    }
]


class ReportBuilderAutomation(BaseAutomation):
    """Natural language report builder with Odoo query execution and export."""

    automation_type = "reporting"
    watched_models = []

    def parse_query(self, query_text: str) -> dict[str, Any]:
        """
        Use Claude to parse a natural language report request into
        a structured Odoo query with model, fields, domain, and group_by.
        """
        today_str = date.today().isoformat()
        system_prompt = REPORT_PARSE_PROMPT.format(today=today_str)

        try:
            result = self.analyze_with_tools(
                system_prompt=system_prompt,
                user_message=f"Parse this report request: {query_text}",
                tools=REPORT_PARSE_TOOLS,
            )

            tool_input = result.get("tool_input", {})
            if not tool_input.get("model"):
                return self._fallback_parse(query_text)

            return {
                "model": tool_input.get("model", "sale.order"),
                "fields": tool_input.get("fields", ["name", "amount_total"]),
                "domain": tool_input.get("domain", []),
                "group_by": tool_input.get("group_by", []),
                "order_by": tool_input.get("order_by", ""),
                "limit": tool_input.get("limit", 100),
                "title": tool_input.get("title", query_text),
                "comparison": tool_input.get("comparison"),
            }

        except Exception as exc:
            logger.warning("claude_parse_failed", error=str(exc))
            return self._fallback_parse(query_text)

    def execute_query(self, parsed_query: dict[str, Any]) -> dict[str, Any]:
        """Execute the parsed query against Odoo and return structured data."""
        model = parsed_query["model"]
        fields = parsed_query["fields"]
        domain = parsed_query.get("domain", [])
        limit = parsed_query.get("limit", 100)

        try:
            domain = self._normalize_domain(domain)

            records = self.fetch_related_records(
                model, domain, fields=fields, limit=limit
            )

            group_by = parsed_query.get("group_by", [])
            if group_by:
                data = self._group_records(records, group_by, fields)
            else:
                data = self._format_records(records, fields)

            comparison = parsed_query.get("comparison")
            if comparison and comparison.get("enabled"):
                comp_domain = self._normalize_domain(comparison.get("compare_domain", []))
                comp_records = self.fetch_related_records(
                    model, comp_domain, fields=fields, limit=limit
                )
                if group_by:
                    comp_data = self._group_records(comp_records, group_by, fields)
                else:
                    comp_data = self._format_records(comp_records, fields)
                data["comparison"] = comp_data
                data["comparison_label"] = comparison.get("compare_label", "Previous Period")

            data["title"] = parsed_query.get("title", "Report")
            data["record_count"] = len(records)
            return data

        except Exception as exc:
            logger.error("query_execution_failed", model=model, error=str(exc))
            return {
                "columns": [],
                "rows": [],
                "title": parsed_query.get("title", "Report"),
                "record_count": 0,
                "error": str(exc),
            }

    def generate_report(self, query_text: str) -> dict[str, Any]:
        """Full pipeline: parse NL query -> execute -> return results."""
        parsed = self.parse_query(query_text)
        result = self.execute_query(parsed)
        return {
            "parsed_query": parsed,
            "result_data": result,
            "status": "completed" if "error" not in result else "error",
            "error_message": result.get("error"),
        }

    def export_excel(self, result_data: dict[str, Any], file_name: str = "") -> str:
        """Export report data to an Excel file using openpyxl."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        os.makedirs(REPORT_DIR, exist_ok=True)
        if not file_name:
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            file_name = f"report_{ts}.xlsx"

        file_path = os.path.join(REPORT_DIR, file_name)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = result_data.get("title", "Report")[:31]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        title = result_data.get("title", "Report")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(result_data.get("columns", [])), 1))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)

        ws.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")

        columns = result_data.get("columns", [])
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=4, column=col_idx, value=col.get("label", col.get("name", "")))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        rows = result_data.get("rows", [])
        for row_idx, row in enumerate(rows, 5):
            for col_idx, col in enumerate(columns, 1):
                col_name = col.get("name", "")
                value = row.get(col_name, "")
                if isinstance(value, (list, tuple)):
                    value = str(value[-1]) if len(value) > 1 else str(value[0]) if value else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

        wb.save(file_path)
        return file_path

    def export_pdf(self, result_data: dict[str, Any], file_name: str = "") -> str:
        """Export report data to a PDF file using basic HTML-to-text approach."""
        os.makedirs(REPORT_DIR, exist_ok=True)
        if not file_name:
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            file_name = f"report_{ts}.txt"

        file_path = os.path.join(REPORT_DIR, file_name)

        title = result_data.get("title", "Report")
        columns = result_data.get("columns", [])
        rows = result_data.get("rows", [])

        lines = [
            f"{'=' * 60}",
            f"  {title}",
            f"  Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            f"{'=' * 60}",
            "",
        ]

        if columns:
            headers = [col.get("label", col.get("name", "")) for col in columns]
            col_widths = [max(len(h), 15) for h in headers]
            header_line = " | ".join(h.ljust(w) for h, w in zip(headers, col_widths))
            lines.append(header_line)
            lines.append("-" * len(header_line))

            for row in rows:
                values = []
                for col, width in zip(columns, col_widths):
                    col_name = col.get("name", "")
                    val = row.get(col_name, "")
                    if isinstance(val, (list, tuple)):
                        val = str(val[-1]) if len(val) > 1 else str(val[0]) if val else ""
                    values.append(str(val).ljust(width))
                lines.append(" | ".join(values))

        lines.append("")
        lines.append(f"Total records: {len(rows)}")

        with open(file_path, "w") as f:
            f.write("\n".join(lines))

        return file_path

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _fallback_parse(query_text: str) -> dict[str, Any]:
        """Simple keyword-based fallback when Claude is unavailable."""
        query_lower = query_text.lower()

        if "sales" in query_lower or "sale" in query_lower:
            model = "sale.order"
            fields = ["name", "partner_id", "amount_total", "state", "date_order"]
        elif "invoice" in query_lower or "billing" in query_lower:
            model = "account.move"
            fields = ["name", "partner_id", "amount_total", "state", "invoice_date"]
        elif "lead" in query_lower or "pipeline" in query_lower or "opportunity" in query_lower:
            model = "crm.lead"
            fields = ["name", "partner_id", "expected_revenue", "probability", "stage_id"]
        elif "product" in query_lower:
            model = "product.template"
            fields = ["name", "list_price", "categ_id", "type"]
        elif "purchase" in query_lower:
            model = "purchase.order"
            fields = ["name", "partner_id", "amount_total", "state", "date_order"]
        elif "expense" in query_lower:
            model = "hr.expense"
            fields = ["name", "employee_id", "total_amount", "state", "date"]
        elif "customer" in query_lower or "contact" in query_lower or "partner" in query_lower:
            model = "res.partner"
            fields = ["name", "email", "phone", "customer_rank"]
        elif "inventory" in query_lower or "stock" in query_lower:
            model = "product.product"
            fields = ["name", "qty_available", "virtual_available", "categ_id"]
        else:
            model = "sale.order"
            fields = ["name", "partner_id", "amount_total", "state", "date_order"]

        domain = []
        today = date.today()

        if "this month" in query_lower:
            start = today.replace(day=1)
            date_field = "date_order" if "sale" in model else "invoice_date" if "account" in model else "create_date"
            domain = [(date_field, ">=", start.isoformat())]
        elif "last month" in query_lower:
            first_this = today.replace(day=1)
            last_month_end = first_this - timedelta(days=1)
            last_month_start = last_month_end.replace(day=1)
            date_field = "date_order" if "sale" in model else "invoice_date" if "account" in model else "create_date"
            domain = [
                (date_field, ">=", last_month_start.isoformat()),
                (date_field, "<=", last_month_end.isoformat()),
            ]
        elif "ytd" in query_lower or "year to date" in query_lower:
            start = today.replace(month=1, day=1)
            date_field = "date_order" if "sale" in model else "invoice_date" if "account" in model else "create_date"
            domain = [(date_field, ">=", start.isoformat())]

        group_by = []
        if "by product" in query_lower or "by category" in query_lower:
            if model == "sale.order":
                model = "sale.order.line"
                fields = ["product_id", "price_subtotal", "product_uom_qty"]
                group_by = ["product_id"]
            elif "category" in query_lower:
                group_by = ["categ_id"]
        elif "by customer" in query_lower or "by partner" in query_lower:
            group_by = ["partner_id"]
        elif "by state" in query_lower or "by status" in query_lower:
            group_by = ["state"]

        return {
            "model": model,
            "fields": fields,
            "domain": domain,
            "group_by": group_by,
            "order_by": "",
            "limit": 100,
            "title": query_text,
            "comparison": None,
        }

    @staticmethod
    def _normalize_domain(domain: list) -> list:
        """Convert domain from JSON (list of lists) to Odoo tuples."""
        if not domain:
            return []
        normalized = []
        for item in domain:
            if isinstance(item, (list, tuple)) and len(item) == 3:
                normalized.append(tuple(item))
            elif isinstance(item, str):
                normalized.append(item)
        return normalized

    @staticmethod
    def _format_records(
        records: list[dict], fields: list[str]
    ) -> dict[str, Any]:
        """Format records into columns + rows structure."""
        columns = [{"name": f, "label": f.replace("_", " ").title()} for f in fields]
        rows = []
        for rec in records:
            row = {}
            for f in fields:
                val = rec.get(f, "")
                row[f] = val
            rows.append(row)
        return {"columns": columns, "rows": rows}

    @staticmethod
    def _group_records(
        records: list[dict], group_by: list[str], fields: list[str]
    ) -> dict[str, Any]:
        """Group records by specified fields and aggregate numeric values."""
        if not group_by or not records:
            return ReportBuilderAutomation._format_records(records, fields)

        group_field = group_by[0]
        groups: dict[str, dict[str, Any]] = {}

        numeric_fields = [
            f for f in fields
            if f != group_field and any(
                records[0].get(f) is not None
                and isinstance(records[0].get(f), (int, float))
                for _ in [0]
            )
        ]

        for rec in records:
            key = rec.get(group_field, "Unknown")
            if isinstance(key, (list, tuple)):
                key = str(key[-1]) if len(key) > 1 else str(key[0]) if key else "Unknown"
            key = str(key)

            if key not in groups:
                groups[key] = {group_field: key, "_count": 0}
                for nf in numeric_fields:
                    groups[key][nf] = 0.0

            groups[key]["_count"] += 1
            for nf in numeric_fields:
                val = rec.get(nf, 0)
                if isinstance(val, (int, float)):
                    groups[key][nf] = groups[key].get(nf, 0) + val

        result_fields = [group_field, "_count"] + numeric_fields
        columns = [
            {"name": f, "label": f.replace("_", " ").title() if f != "_count" else "Count"}
            for f in result_fields
        ]

        rows = list(groups.values())
        return {"columns": columns, "rows": rows}

    @staticmethod
    def _parse_cron_to_description(cron: str) -> str:
        """Convert cron expression to human-readable schedule."""
        parts = cron.split()
        if len(parts) < 5:
            return f"Custom schedule: {cron}"

        minute, hour, dom, month, dow = parts[:5]
        day_names = {"0": "Sunday", "1": "Monday", "2": "Tuesday", "3": "Wednesday",
                     "4": "Thursday", "5": "Friday", "6": "Saturday", "7": "Sunday",
                     "MON": "Monday", "TUE": "Tuesday", "WED": "Wednesday",
                     "THU": "Thursday", "FRI": "Friday", "SAT": "Saturday", "SUN": "Sunday"}

        if dow != "*":
            day = day_names.get(dow.upper(), dow)
            return f"Every {day} at {hour}:{minute.zfill(2)}"
        if dom != "*":
            return f"Monthly on day {dom} at {hour}:{minute.zfill(2)}"
        return f"Daily at {hour}:{minute.zfill(2)}"
